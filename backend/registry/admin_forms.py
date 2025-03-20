import openpyxl

from django import forms
from django.core.exceptions import ValidationError

from . import models


class OrganisationAdminForm(forms.ModelForm):
    class Meta:
        model = models.Organisation
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        sites = set(cleaned_data.get("sites", models.Site.objects.none()).values_list("id", flat=True))
        people = cleaned_data.get("people", [])

        for person in people:
            person_sites = set(person.sites.values_list("id", flat=True))
            if not person_sites.intersection(sites):
                self.add_error('people', f"Person {person} is not on the same site as the organisation.")

        return cleaned_data


class PersonImportForm(forms.Form):
    file = forms.FileField()

    def clean_file(self):
        file = self.cleaned_data.get('file')
        if not file:
            raise ValidationError("No file uploaded")

        if not file.name.endswith('.xlsx'):
            raise ValidationError("File must be an Excel file (.xlsx)")

        try:
            # Try to read both sheets to validate
            workbook = openpyxl.load_workbook(file, read_only=True)

            # Validate 'Organisations' sheet
            if 'Organisations' not in workbook.sheetnames:
                raise ValidationError("Excel file must contain an 'Organisations' sheet")

            # Validate 'People' sheet
            if 'People' not in workbook.sheetnames:
                raise ValidationError("Excel file must contain a 'People' sheet")

            # Validate columns in 'Organisations' sheet
            org_sheet = workbook['Organisations']
            org_headers = [str(cell.value).strip() if cell.value else '' for cell in org_sheet[1]]  # First row
            required_org_headers = ['Short name', 'Country', 'Name']
            missing_org_headers = [header for header in required_org_headers if header not in org_headers]
            if missing_org_headers:
                raise ValidationError(f"Organisations is sheet missing required columns: {', '.join(missing_org_headers)}")

            # Check for unexpected headers in Organisations sheet
            unexpected_org_headers = [header for header in org_headers if header and header not in required_org_headers]
            if unexpected_org_headers:
                raise ValidationError(f"Unexpected columns in Organisations sheet: {', '.join(unexpected_org_headers)}")

            # Validate columns in 'People' sheet
            people_sheet = workbook['People']
            people_headers = [str(cell.value).strip() if cell.value else '' for cell in people_sheet[1]]  # First row
            required_people_headers = ['ORCID', 'Organisation (short)', 'First name', 'Last name', 'Email']
            missing_people_headers = [header for header in required_people_headers if header not in people_headers]
            if missing_people_headers:
                raise ValidationError(f"People sheet is missing required columns: {', '.join(missing_people_headers)}")

            # Check for unexpected headers in People sheet (excluding the optional ORCID)
            unexpected_people_headers = [header for header in people_headers if header and header not in required_people_headers]
            if unexpected_people_headers:
                raise ValidationError(f"Unexpected columns in People sheet: {', '.join(unexpected_people_headers)}")

            file.seek(0)  # Reset file pointer after reading
        except ValidationError:
            raise
        except Exception as e:
            raise ValidationError(f"Error reading Excel file: {str(e)}")

        return file

    def add_warning(self, message):
        # Store warnings to be displayed later
        if not hasattr(self, 'warnings'):
            self.warnings = []
        self.warnings.append(message)


class GroupAdminForm(forms.ModelForm):
    class Meta:
        model = models.Group
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        site = cleaned_data.get("site", None)

        if site is not None:
            people = cleaned_data.get("people", [])
            for person in people:
                if site not in person.sites.all():
                    self.add_error('people', f"Person {person} is not on the same site as the group.")

        return cleaned_data


class ResourceAdminForm(forms.ModelForm):
    class Meta:
        model = models.Resource
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        site = cleaned_data.get("site", None)

        if site is not None:
            status = cleaned_data.get("status")
            if status is not None and status.site != site:
                self.add_error('status', "Status is not on the same site as the resource.")

            license = cleaned_data.get("license")
            if license is not None and site not in license.sites.all():
                self.add_error('license', "License is not on the same site as the resource.")

            for group in cleaned_data.get("groups", models.Group.objects.none()).select_related("site"):
                if site != group.site:
                    self.add_error('groups', f"Group {group} is not on the same site as the resource.")

        return cleaned_data


class StudyDesignAdminForm(forms.ModelForm):
    class Meta:
        model = models.StudyDesign
        fields = '__all__'

    def clean(self):
        cleaned_data = super().clean()
        site = cleaned_data.get("site", None)

        if site is not None:
            license = cleaned_data.get("license")
            if license is not None and site not in license.sites.all():
                self.add_error('license', "License is not on the same site as the resource.")

            for group in cleaned_data.get("groups", models.Group.objects.none()).select_related("site"):
                if site != group.site:
                    self.add_error('groups', f"Group {group} is not on the same site as the resource.")

        return cleaned_data
